"""
D-GITALCODE ExtractorX - Report service.

Product: D-GITALCODE ExtractorX | https://dgitalcode.ma

Generates branded TXT, PDF and XLSX extraction reports.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable, List, Optional

import config
from core.extractor import ExtractionSummary
from utils.format_utils import format_duration, format_size, format_timestamp

_SEPARATOR = "=" * 80
_SUBSEPARATOR = "-" * 80


class ReportService:
    """Produce extraction reports in TXT, PDF, XLSX and metadata formats.

    Reports are generated per source document, inside that document's own
    output folder, so every extracted file is fully self-contained.
    """

    def __init__(self, logger: logging.Logger) -> None:
        self._logger = logger

    def generate_all(
        self,
        summary: ExtractionSummary,
        output_dir: Path,
        formats: Optional[Iterable[str]] = None,
    ) -> List[Path]:
        """Generate the selected report formats (default: all) in ``output_dir``.

        Args:
            summary: Statistics scoped to a single source document.
            output_dir: The document's dedicated output folder.
            formats: Subset of {"txt", "pdf", "xlsx", "metadata"}.

        PDF/XLSX generation degrades gracefully when the optional
        dependencies (reportlab / openpyxl) are unavailable.
        """
        generators = {
            "txt": self.generate_txt,
            "pdf": self.generate_pdf,
            "xlsx": self.generate_xlsx,
            "metadata": self.generate_metadata_xlsx,
        }
        selected = list(formats) if formats is not None else list(generators)

        generated: List[Path] = []
        for name in selected:
            generator = generators.get(name)
            if generator is None:
                self._logger.warning("Unknown report format: %s", name)
                continue
            try:
                path = generator(summary, output_dir)
                if path is not None:
                    generated.append(path)
                    self._logger.info("Report generated: %s", path.name)
            except Exception as exc:
                self._logger.error("Report generation failed (%s): %s",
                                   generator.__name__, exc)
        return generated

    # ----------------------------------------------------------------- TXT

    def generate_txt(self, summary: ExtractionSummary, output_dir: Path) -> Path:
        """Write the plain-text summary report."""
        path = output_dir / config.REPORT_TXT_FILENAME
        breakdown = summary.format_breakdown

        source_name = (
            summary.results[0].source.name if len(summary.results) == 1
            else f"{len(summary.results)} document(s)"
        )
        lines: List[str] = [
            _SEPARATOR,
            f"{config.BRAND} IMAGE EXTRACTION REPORT",
            _SEPARATOR,
            "",
            f"Brand:       {config.BRAND}",
            f"Brand:       {config.BRAND}",
            f"Application: {config.APP_NAME} v{config.APP_VERSION}",
            "",
            "EXTRACTION DETAILS:",
            _SUBSEPARATOR,
            f"Source Document:  {source_name}",
            f"Timestamp:        {format_timestamp(summary.started_at)}",
            f"Processing Time:  {format_duration(summary.processing_time_seconds)}",
            f"Total Data:       {format_size(summary.total_size_bytes)}",
            "",
            "IMAGE STATISTICS:",
            _SUBSEPARATOR,
            f"Total Images Extracted: {summary.total_images}",
            f"Duplicates Detected:    {summary.duplicate_count}",
            "",
            "Images by Format:",
        ]
        lines += [
            f"  {fmt:8s}: {count:4d} image(s)"
            for fmt, count in breakdown.items() if count > 0
        ]
        lines += [
            "",
            "SOURCE DOCUMENT:",
            _SUBSEPARATOR,
        ]
        for result in summary.results:
            status = "OK " if result.success else "FAIL"
            lines.append(
                f"  [{status}] {result.source.name} "
                f"({result.total_images} images, "
                f"{result.duplicate_count} duplicates)"
            )
            lines.append(f"         Path:    {result.source.resolve()}")
            if result.success and result.format_breakdown:
                formats = ", ".join(
                    f"{fmt}: {count}"
                    for fmt, count in sorted(result.format_breakdown.items())
                )
                lines.append(f"         Formats: {formats}")
            if result.error:
                lines.append(f"         Error: {result.error}")
        lines += [
            "",
            _SEPARATOR,
            f"Output Folder: {output_dir.resolve()}",
            _SEPARATOR,
            "END OF REPORT",
            _SEPARATOR,
        ]
        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    # ----------------------------------------------------------------- PDF

    def generate_pdf(self, summary: ExtractionSummary, output_dir: Path) -> Path | None:
        """Write the branded PDF report (requires reportlab)."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.pdfgen import canvas
        except ImportError:
            self._logger.warning("reportlab not installed - skipping PDF report")
            return None

        path = output_dir / config.REPORT_PDF_FILENAME
        pdf = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4

        # Watermark
        pdf.saveState()
        pdf.setFont("Helvetica-Bold", 60)
        pdf.setFillColor(colors.Color(0, 0.66, 0.91, alpha=0.08))
        pdf.translate(width / 2, height / 2)
        pdf.rotate(45)
        pdf.drawCentredString(0, 0, config.BRAND)
        pdf.restoreState()

        # Header band
        pdf.setFillColor(colors.HexColor("#00a8e8"))
        pdf.rect(0, height - 3 * cm, width, 3 * cm, fill=True, stroke=False)
        pdf.setFillColor(colors.white)
        pdf.setFont("Helvetica-Bold", 20)
        pdf.drawString(2 * cm, height - 1.7 * cm, f"{config.BRAND} Extraction Report")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(2 * cm, height - 2.4 * cm,
                       f"{config.BRAND}  |  {config.APP_NAME} v{config.APP_VERSION}")

        y = height - 4.5 * cm

        def write_line(text: str, bold: bool = False, indent: float = 0.0) -> None:
            nonlocal y
            if y < 3 * cm:
                pdf.showPage()
                y = height - 3 * cm
            pdf.setFillColor(colors.black)
            pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 11)
            pdf.drawString(2 * cm + indent, y, text)
            y -= 0.65 * cm

        write_line("Extraction Summary", bold=True)
        if len(summary.results) == 1:
            write_line(f"Source Document: {summary.results[0].source.name}",
                       indent=0.5 * cm)
        write_line(f"Timestamp: {format_timestamp(summary.started_at)}", indent=0.5 * cm)
        write_line(f"Total Images: {summary.total_images}", indent=0.5 * cm)
        write_line(f"Duplicates Detected: {summary.duplicate_count}", indent=0.5 * cm)
        write_line(f"Processing Time: "
                   f"{format_duration(summary.processing_time_seconds)}",
                   indent=0.5 * cm)
        write_line(f"Total Data: {format_size(summary.total_size_bytes)}",
                   indent=0.5 * cm)

        y -= 0.3 * cm
        write_line("Format Breakdown", bold=True)
        for fmt, count in summary.format_breakdown.items():
            if count > 0:
                write_line(f"{fmt}: {count} image(s)", indent=0.5 * cm)

        y -= 0.3 * cm
        write_line("Source Document", bold=True)
        for result in summary.results:
            status = "OK" if result.success else "FAILED"
            write_line(
                f"[{status}] {result.source.name} - {result.total_images} images",
                indent=0.5 * cm,
            )

        # Footer
        pdf.setFont("Helvetica-Oblique", 8)
        pdf.setFillColor(colors.grey)
        pdf.drawCentredString(width / 2, 1.5 * cm, config.COPYRIGHT)
        pdf.save()
        return path

    # ---------------------------------------------------------------- XLSX

    def generate_xlsx(self, summary: ExtractionSummary, output_dir: Path) -> Path | None:
        """Write the structured Excel report (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            self._logger.warning("openpyxl not installed - skipping XLSX report")
            return None

        path = output_dir / config.REPORT_XLSX_FILENAME
        workbook = Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="00A8E8")
        title_font = Font(bold=True, size=14, color="00A8E8")

        # --- Summary sheet
        sheet = workbook.active
        sheet.title = "Summary"
        sheet["A1"] = f"{config.BRAND} Extraction Report"
        sheet["A1"].font = title_font
        sheet["A2"] = f"{config.BRAND} — {config.COMPANY_SITE}"
        sheet["A3"] = f"Application: {config.APP_NAME} v{config.APP_VERSION}"

        source_name = (
            summary.results[0].source.name if len(summary.results) == 1
            else f"{len(summary.results)} document(s)"
        )
        rows = [
            ("Source Document", source_name),
            ("Timestamp", format_timestamp(summary.started_at)),
            ("Total Images", summary.total_images),
            ("Duplicates Detected", summary.duplicate_count),
            ("Processing Time", format_duration(summary.processing_time_seconds)),
            ("Total Data", format_size(summary.total_size_bytes)),
        ]
        for offset, (label, value) in enumerate(rows, start=5):
            sheet.cell(row=offset, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=offset, column=2, value=value)
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 40

        # --- Format breakdown sheet
        formats_sheet = workbook.create_sheet("Formats")
        for col, header in enumerate(("Format", "Count"), start=1):
            cell = formats_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row, (fmt, count) in enumerate(summary.format_breakdown.items(), start=2):
            formats_sheet.cell(row=row, column=1, value=fmt)
            formats_sheet.cell(row=row, column=2, value=count)
        formats_sheet.column_dimensions["A"].width = 14

        # --- Images detail sheet
        images_sheet = workbook.create_sheet("Images")
        headers = ("File", "Format", "Size (bytes)", "SHA-256",
                   "Duplicate", "Source Document")
        for col, header in enumerate(headers, start=1):
            cell = images_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row = 2
        for result in summary.results:
            for image in result.images:
                images_sheet.cell(
                    row=row, column=1,
                    value=image.saved_path.name if image.saved_path else "(skipped)",
                )
                images_sheet.cell(row=row, column=2, value=image.format_folder)
                images_sheet.cell(row=row, column=3, value=image.size_bytes)
                images_sheet.cell(row=row, column=4, value=image.sha256)
                images_sheet.cell(row=row, column=5,
                                  value="Yes" if image.is_duplicate else "No")
                images_sheet.cell(row=row, column=6, value=image.source_document)
                row += 1
        for letter, width in (("A", 20), ("B", 10), ("C", 14), ("D", 68),
                              ("E", 10), ("F", 32)):
            images_sheet.column_dimensions[letter].width = width

        workbook.save(path)
        return path

    # ------------------------------------------------------------ metadata

    def generate_metadata_xlsx(
        self, summary: ExtractionSummary, output_dir: Path
    ) -> Path | None:
        """Write ``Image_Metadata.xlsx`` with full per-image metadata."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self._logger.warning("openpyxl not installed - skipping metadata export")
            return None

        path = output_dir / config.REPORT_METADATA_FILENAME
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Image Metadata"

        sheet["A1"] = f"{config.BRAND} Image Metadata Export"
        sheet["A1"].font = Font(bold=True, size=14, color="00A8E8")
        sheet["A2"] = (
            f"{config.BRAND} | {config.APP_NAME} v{config.APP_VERSION} | "
            f"{format_timestamp(summary.started_at)}"
        )

        headers = (
            "Source File", "Image Name", "Format", "Width", "Height",
            "Size (KB)", "Hash (SHA-256)", "Output Path",
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="00A8E8")
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 5
        for result in summary.results:
            for image in result.images:
                sheet.cell(row=row, column=1, value=image.source_document)
                sheet.cell(
                    row=row, column=2,
                    value=image.saved_path.name if image.saved_path else "(skipped)",
                )
                sheet.cell(row=row, column=3, value=image.format_folder)
                sheet.cell(row=row, column=4, value=image.width)
                sheet.cell(row=row, column=5, value=image.height)
                sheet.cell(
                    row=row, column=6, value=round(image.size_bytes / 1024, 2)
                )
                sheet.cell(row=row, column=7, value=image.sha256)
                sheet.cell(
                    row=row, column=8,
                    value=str(image.saved_path) if image.saved_path else "",
                )
                row += 1

        for letter, width in (("A", 30), ("B", 18), ("C", 9), ("D", 8),
                              ("E", 8), ("F", 10), ("G", 68), ("H", 60)):
            sheet.column_dimensions[letter].width = width

        workbook.save(path)
        return path
